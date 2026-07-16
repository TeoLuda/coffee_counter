from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime

# -------------------------
# SUPABASE
# -------------------------

SUPABASE_URL = "https://vybtsrmxrhqlplpaotkq.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InZ5YnRzcm14cmhxbHBscGFvdGtxIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzkyNzA5MDMsImV4cCI6MjA5NDg0NjkwM30.v1JZ06whw0CvgchJRC62s0fbG-21Yp3qvLLp_eoIiVc"

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# -------------------------
# SETTINGS
# -------------------------

PRICE_PER_COFFEE = 0.50

# -------------------------
# DOWNLOAD DATA
# -------------------------

response = (
    supabase
    .table("users")
    .select("name, email, coffee_count")
    .order("name")
    .execute()
)

users = response.data

# -------------------------
# CREATE EXCEL
# -------------------------

wb = Workbook()
ws = wb.active

ws.title = "Coffee Report"

headers = [
    "Name",
    "Email",
    "Coffees",
    "Price per coffee (€)",
    "Total (€)"
]

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True)

total_coffees = 0
total_money = 0

row = 2

for user in users:

    name = user["name"]
    email = user.get("email", "")
    coffees = user["coffee_count"]

    amount = coffees * PRICE_PER_COFFEE

    ws.cell(row=row, column=1).value = name
    ws.cell(row=row, column=2).value = email
    ws.cell(row=row, column=3).value = coffees
    ws.cell(row=row, column=4).value = PRICE_PER_COFFEE
    ws.cell(row=row, column=5).value = amount

    total_coffees += coffees
    total_money += amount

    row += 1

# Totals

row += 1

ws.cell(row=row, column=1).value = "TOTAL"
ws.cell(row=row, column=1).font = Font(bold=True)

ws.cell(row=row, column=3).value = total_coffees
ws.cell(row=row, column=5).value = total_money

# Autosize columns

for column in ws.columns:

    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)

    ws.column_dimensions[get_column_letter(column[0].column)].width = length + 3

# Save

today = datetime.now().strftime("%Y-%m-%d")

filename = f"Coffee_Report_{today}.xlsx"

wb.save(filename)

print()
print("Report created:")
print(filename)
print(f"Total coffees: {total_coffees}")
print(f"Total money: €{total_money:.2f}")
